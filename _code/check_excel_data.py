# -*- coding: utf-8 -*-
"""Excel 파일 실제 데이터 확인"""

import openpyxl

excel_path = "M:/CALife/CAShop - 구매대행/01.All Product List.xlsx"

from pathlib import Path

excel_file = Path(excel_path)

print("="*80)
print("📊 Excel File Analysis")
print("="*80)
print(f"File: {excel_path}")
print(f"Exists: {'✅ Yes' if excel_file.exists() else '❌ No'}")

if not excel_file.exists():
    print(f"\n❌ 파일이 존재하지 않습니다!")
    print("="*80)
    exit(1)

print(f"Size: {excel_file.stat().st_size:,} bytes")
print(f"Modified: {excel_file.stat().st_mtime}")
print()

# Excel 파일 열기
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

# 모든 시트 정보
print("="*80)
print("📋 All Sheets")
print("="*80)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"{i}. {sheet_name}")
    print(f"   - Max Row: {ws.max_row}")
    print(f"   - Max Column: {ws.max_column}")
    
    # 실제 데이터가 있는지 확인 (첫 10행)
    has_data = False
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=ws.max_column):
        if any(cell.value is not None and str(cell.value).strip() for cell in row):
            has_data = True
            break
    
    print(f"   - Has Data: {'✅ Yes' if has_data else '❌ No (empty)'}")
    print()

# Purchase 시트 상세 확인
print("="*80)
print("🔍 Purchase Sheet Detail")
print("="*80)

ws_purchase = wb['Purchase']
print(f"Max Row: {ws_purchase.max_row}")
print(f"Max Column: {ws_purchase.max_column}")
print()

print("First 10 rows:")
for i, row in enumerate(ws_purchase.iter_rows(min_row=1, max_row=10, max_col=15), 1):
    values = [cell.value for cell in row]
    if any(v is not None for v in values):
        print(f"Row {i}: {values}")

# Sheet1도 확인
if 'Sheet1' in wb.sheetnames:
    print("\n" + "="*80)
    print("🔍 Sheet1 Detail")
    print("="*80)
    
    ws_sheet1 = wb['Sheet1']
    print(f"Max Row: {ws_sheet1.max_row}")
    print(f"Max Column: {ws_sheet1.max_column}")
    print()
    
    print("First 10 rows:")
    for i, row in enumerate(ws_sheet1.iter_rows(min_row=1, max_row=10, max_col=15), 1):
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            print(f"Row {i}: {values}")

wb.close()

print("\n" + "="*80)
print("✅ Analysis Complete")
print("="*80)
